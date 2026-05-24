"""Tests for Excel export and the keyless web/weather tools."""

from __future__ import annotations

import httpx
import respx
from openpyxl import load_workbook

from aegis.assistant.web_tools import get_weather, web_search
from aegis.storage.store import Storage


def test_save_xlsx_creates_readable_workbook():
    storage = Storage(run_name="xlsx-test")
    records = [
        {"url": "https://a.com", "ok": True, "title": "Alpha"},
        {"url": "https://b.com", "ok": False, "error": "404"},
    ]
    path = storage.save_xlsx(records, name="results")
    assert path.exists()

    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    # Header is the union of keys across both records.
    assert "url" in header and "title" in header and "error" in header
    # Two data rows below the header.
    assert ws.max_row == 3
    # Header row is bold (formatting applied).
    assert ws["A1"].font.bold is True


def test_save_xlsx_empty_records():
    storage = Storage(run_name="empty")
    path = storage.save_xlsx([], name="empty")
    assert path.exists()


def test_save_xlsx_nested_values_are_json_encoded():
    storage = Storage(run_name="nested")
    records = [{"url": "https://a.com", "extracted": {"k": "v"}}]
    path = storage.save_xlsx(records, name="n")
    wb = load_workbook(path)
    ws = wb.active
    # The nested dict should appear as a JSON string in its cell.
    values = [c.value for c in ws[2]]
    assert any(isinstance(v, str) and '"k"' in v for v in values)


@respx.mock
def test_web_search_returns_hits():
    html = """
    <div class="result">
      <a class="result__a" href="https://example.com">Example</a>
      <div class="result__snippet">An example result.</div>
    </div>
    """
    respx.post("https://html.duckduckgo.com/html/").mock(
        return_value=httpx.Response(200, text=html)
    )
    hits = web_search("example query")
    assert len(hits) == 1
    assert hits[0].title == "Example"
    assert hits[0].url == "https://example.com"


@respx.mock
def test_web_search_offline_returns_empty():
    respx.post("https://html.duckduckgo.com/html/").mock(
        side_effect=httpx.ConnectError("offline")
    )
    assert web_search("anything") == []


@respx.mock
def test_weather_handles_unknown_place():
    respx.get("https://geocoding-api.open-meteo.com/v1/search").mock(
        return_value=httpx.Response(200, json={"results": []})
    )
    out = get_weather("Nowheresville")
    assert "error" in out


@respx.mock
def test_weather_success():
    respx.get("https://geocoding-api.open-meteo.com/v1/search").mock(
        return_value=httpx.Response(
            200,
            json={"results": [{"name": "Paris", "country": "France",
                               "latitude": 48.85, "longitude": 2.35}]},
        )
    )
    respx.get("https://api.open-meteo.com/v1/forecast").mock(
        return_value=httpx.Response(
            200,
            json={"current": {"temperature_2m": 18.0, "relative_humidity_2m": 60,
                              "wind_speed_10m": 12.0}},
        )
    )
    out = get_weather("Paris")
    assert out["location"] == "Paris, France"
    assert out["temperature_c"] == 18.0
