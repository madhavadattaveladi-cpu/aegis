"""Local file management for scraped data.

Persists scraped records to the output directory in a few formats and keeps a
tidy, timestamped folder structure. This is the "local file management" piece
of the pipeline.
"""

from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from aegis.utils.config import get_settings
from aegis.utils.logging import get_logger

log = get_logger(__name__)


def _slugify(text: str, max_len: int = 40) -> str:
    """Make a filesystem-safe slug from arbitrary text."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return (text or "untitled")[:max_len]


class Storage:
    """Writes scraped data to disk under a per-run timestamped folder."""

    def __init__(self, run_name: str = "scrape") -> None:
        settings = get_settings()
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        self.run_dir = settings.output_dir / f"{_slugify(run_name)}-{stamp}"
        self.run_dir.mkdir(parents=True, exist_ok=True)
        log.info("Storage run directory: %s", self.run_dir)

    def save_json(self, records: list[dict[str, Any]], name: str = "results") -> Path:
        """Save all records as a single pretty-printed JSON array."""
        path = self.run_dir / f"{_slugify(name)}.json"
        path.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
        return path

    def save_jsonl(self, records: Iterable[dict[str, Any]], name: str = "results") -> Path:
        """Save records as JSON Lines (one object per line) — great for big data."""
        path = self.run_dir / f"{_slugify(name)}.jsonl"
        with path.open("w", encoding="utf-8") as fh:
            for rec in records:
                fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
        return path

    def save_csv(self, records: list[dict[str, Any]], name: str = "results") -> Path:
        """Save records as CSV using the union of all keys as headers."""
        path = self.run_dir / f"{_slugify(name)}.csv"
        if not records:
            path.write_text("", encoding="utf-8")
            return path
        fieldnames: list[str] = []
        for rec in records:
            for key in rec:
                if key not in fieldnames:
                    fieldnames.append(key)
        with path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            for rec in records:
                writer.writerow(rec)
        return path

    def save_text(self, content: str, name: str) -> Path:
        """Save a single text/markdown file (e.g. an AI-written summary)."""
        path = self.run_dir / name
        path.write_text(content, encoding="utf-8")
        return path

    def save_xlsx(self, records: list[dict[str, Any]], name: str = "results") -> Path:
        """Save records to a formatted .xlsx workbook.

        Uses openpyxl (a project dependency). Only flat, scalar values are
        written; nested dicts/lists are JSON-encoded into a single cell so the
        sheet stays rectangular. A bold, shaded header row and auto-ish column
        widths give a professional look.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        path = self.run_dir / f"{_slugify(name)}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = _slugify(name)[:31] or "results"  # Excel sheet-name limit

        if not records:
            wb.save(path)
            return path

        # Header = union of keys, preserving first-seen order.
        headers: list[str] = []
        for rec in records:
            for key in rec:
                if key not in headers:
                    headers.append(key)

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="305496")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows.
        for rec in records:
            row: list[Any] = []
            for key in headers:
                value = rec.get(key, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                row.append(value)
            ws.append(row)

        # Reasonable column widths capped so long text doesn't explode the sheet.
        for col_idx, header in enumerate(headers, start=1):
            longest = len(str(header))
            for rec in records:
                longest = max(longest, len(str(rec.get(header, ""))))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max(longest + 2, 10), 60
            )

        ws.freeze_panes = "A2"  # keep header visible when scrolling
        wb.save(path)
        return path
